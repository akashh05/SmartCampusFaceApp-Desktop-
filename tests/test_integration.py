import unittest
import os
import database
import auth
import face_recognition
from openpyxl import load_workbook

class TestIntegrationFaceRecognitionApp(unittest.TestCase):

    def setUp(self):
        """Prepare test admin and student for integration test"""
        self.admin_username = "integrationadmin"
        self.admin_password = "admin123"
        self.student = {
            "name": "Integration Test Student",
            "contact": "9998887777",
            "roll_number": "INT001",
            "face_image": "int001.jpg"
        }

        # Add admin (ignore if exists)
        database.add_admin(self.admin_username, self.admin_password)

        # Add student
        database.add_student(**self.student)

    def test_admin_login_and_student_workflow(self):
        """Test full flow: Admin login → Register student → Log behavior"""
        
        # Step 1: Authenticate Admin
        login_result = auth.authenticate(self.admin_username, self.admin_password)
        self.assertTrue(login_result)

        # Step 2: Get student from DB
        fetched = database.get_student_by_roll(self.student["roll_number"])
        self.assertIsNotNone(fetched)
        self.assertEqual(fetched["name"], self.student["name"])

        # Step 3: Train recognizer
        face_recognition.train_faces()
        self.assertTrue(os.path.exists("dataset/trainer.yml"))

        # Step 4: Log misbehavior
        face_recognition.log_misbehavior(self.student, reason="Integration Testing")
        self.assertTrue(os.path.exists("misbehavior_log.xlsx"))

        # Step 5: Confirm log is saved
        wb = load_workbook("misbehavior_log.xlsx")
        ws = wb.active
        found = any(row[1] == self.student["roll_number"] for row in ws.iter_rows(min_row=2, values_only=True))
        self.assertTrue(found)

    def tearDown(self):
        """Clean up if needed (optional for test DB)"""
        # You can delete test student or clear logs here if using test DB
        pass


if __name__ == "__main__":
    unittest.main()
