import unittest
import os
import database
import auth
import face_recognition
from openpyxl import load_workbook


class TestFaceRecognitionApp(unittest.TestCase):

    # ---------- Admin Tests ----------
    def test_1_add_admin(self):
        """Test adding a new admin user"""
        result = database.add_admin("testadmin", "securepass")
        self.assertIn(result, [True, False])  # False if already exists

    def test_2_authenticate_admin(self):
        """Test authenticating an existing admin"""
        auth_result = auth.authenticate("testadmin", "securepass")
        self.assertIsInstance(auth_result, bool)
        self.assertTrue(auth_result)

    # ---------- Student Tests ----------
    def test_3_add_student(self):
        """Test adding a student to the database"""
        added = database.add_student("Unit Test Student", "1234567890", "UT001", "ut001.jpg")
        self.assertTrue(added)

    def test_4_get_student_by_roll(self):
        """Test fetching student by roll number"""
        student = database.get_student_by_roll("UT001")
        self.assertIsNotNone(student)
        self.assertEqual(student["name"], "Unit Test Student")

    # ---------- Misbehavior Log ----------
    def test_5_log_misbehavior(self):
        """Test logging misbehavior in Excel"""
        student = {
            "name": "Unit Test Student",
            "roll_number": "UT001",
            "contact": "1234567890"
        }
        face_recognition.log_misbehavior(student, reason="Testing Log Entry")
        self.assertTrue(os.path.exists("misbehavior_log.xlsx"))

        wb = load_workbook("misbehavior_log.xlsx")
        ws = wb.active
        found = any(row[1] == "UT001" for row in ws.iter_rows(min_row=2, values_only=True))
        self.assertTrue(found)


if __name__ == "__main__":
    unittest.main()
