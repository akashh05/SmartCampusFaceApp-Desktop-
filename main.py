import customtkinter as ctk
from tkinter import messagebox
from tkinter.filedialog import asksaveasfilename
from PIL import Image
import csv
import cv2
import os
import database
import face_recognition as fr
import auth
import openpyxl

# print(cv2.__version__)

# ✅ Fixed cascade load using OpenCV's built-in path
CASCADE_PATH = os.path.join(cv2.data.haarcascades, 'haarcascade_frontalface_default.xml')
face_cascade = cv2.CascadeClassifier(CASCADE_PATH)

# Ensure recognizer is available
try:
    recognizer = cv2.face.LBPHFaceRecognizer_create()
except AttributeError:
    messagebox.showerror("OpenCV Error", "LBPH recognizer not found. Make sure opencv-contrib-python is installed.")
    exit()

# GUI config
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("🎭 Face Recognition System")
app.geometry("750x850")

# Background
try:
    bg_image = ctk.CTkImage(light_image=Image.open("assets/bg.jpg"), size=(750, 850))
    bg_label = ctk.CTkLabel(app, image=bg_image)
    bg_label.place(relwidth=1, relheight=1)
except Exception as e:
    messagebox.showerror("Image Error", f"Background image error: {e}")

# Styles
button_style = {
    "width": 350,
    "height": 60,
    "corner_radius": 25,
    "font": ("Arial", 20, "bold")
}
hover_color = "#1E90FF"

# -------- Features --------
def register_student():
    win = ctk.CTkToplevel(app)
    win.title("Register Student")
    win.geometry("450x550")

    ctk.CTkLabel(win, text="Name:", font=("Arial", 18)).pack(pady=5)
    name_entry = ctk.CTkEntry(win, width=350, height=50, font=("Arial", 16))
    name_entry.pack()

    ctk.CTkLabel(win, text="Contact:", font=("Arial", 18)).pack(pady=5)
    contact_entry = ctk.CTkEntry(win, width=350, height=50, font=("Arial", 16))
    contact_entry.pack()

    ctk.CTkLabel(win, text="Roll Number:", font=("Arial", 18)).pack(pady=5)
    roll_entry = ctk.CTkEntry(win, width=350, height=50, font=("Arial", 16))
    roll_entry.pack()

    def capture():
        fr.capture_face(name_entry.get(), contact_entry.get(), roll_entry.get())

    ctk.CTkButton(win, text="📸 Capture Face", command=capture,
                  fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=30)

def train_faces():
    fr.train_faces()
    messagebox.showinfo("✅ Training Complete", "Model trained successfully!")

def recognize_student():
    student = fr.recognize_face()
    if not student:
        messagebox.showerror("❌ Not Found", "Face not recognized!")
        return

    # Fetch misbehavior count
    roll = student["roll_number"]
    count = 0
    history = []

    try:
        wb = openpyxl.load_workbook("misbehavior_log.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[1]) == str(roll):
                count += 1
                history.append(row)
    except FileNotFoundError:
        pass  # No logs found

    # Show student info + misbehavior count
    info_text = f"Name: {student['name']}\nContact: {student['contact']}\nRoll: {roll}\n🚨 Past Misbehavior Count: {count}"
    result = messagebox.askyesno("🎉 Match Found", info_text + "\n\nDo you want to view incident history?")
    
    if result and count > 0:
        view_student_history(roll)

    # Ask to log new misbehavior
    if messagebox.askyesno("⚠️ Misbehavior", f"Log new misbehavior for {student['name']}?"):
        reason = ctk.CTkInputDialog(text="Enter reason for misbehavior:", title="Misbehavior Log").get_input()
        if reason:
            fr.log_misbehavior(student, reason)
            messagebox.showinfo("✅ Logged", "Misbehavior logged successfully.")

def view_student_history(roll_number):
    try:
        wb = openpyxl.load_workbook("misbehavior_log.xlsx")
        ws = wb.active
    except FileNotFoundError:
        messagebox.showinfo("Not Found", "No log file found.")
        return

    student_logs = [row for row in ws.iter_rows(min_row=2, values_only=True) if str(row[1]) == str(roll_number)]

    if not student_logs:
        messagebox.showinfo("Info", f"No logs found for Roll Number: {roll_number}")
        return

    win = ctk.CTkToplevel(app)
    win.title(f"📄 Misbehavior History - Roll: {roll_number}")
    win.geometry("700x400")

    ctk.CTkLabel(win, text=f"📄 Incident History for Roll No: {roll_number}", font=("Arial", 20, "bold")).pack(pady=10)

    frame = ctk.CTkScrollableFrame(win, width=650, height=300)
    frame.pack(pady=10)

    headers = ["Name", "Roll", "Contact", "Date", "Time", "Reason"]
    ctk.CTkLabel(frame, text=" | ".join(headers), font=("Arial", 16, "bold")).pack(anchor="w", padx=10, pady=3)

    for row in student_logs:
        ctk.CTkLabel(frame, text=" | ".join(str(cell) for cell in row), font=("Arial", 14)).pack(anchor="w", padx=10)

def view_misbehavior_logs():
    try:
        wb = openpyxl.load_workbook("misbehavior_log.xlsx")
        ws = wb.active
    except FileNotFoundError:
        messagebox.showinfo("Not Found", "No misbehavior log found.")
        return

    win = ctk.CTkToplevel(app)
    win.title("🚨 Misbehavior Logs")
    win.geometry("700x500")
    ctk.CTkLabel(win, text="🚨 Misbehavior Incidents", font=("Arial", 22, "bold")).pack(pady=10)

    log_frame = ctk.CTkScrollableFrame(win, width=650, height=400)
    log_frame.pack(pady=10)

    headers = [cell.value for cell in ws[1]]
    ctk.CTkLabel(log_frame, text=" | ".join(headers), font=("Arial", 16, "bold")).pack(anchor="w", padx=10, pady=5)

    for row in ws.iter_rows(min_row=2, values_only=True):
        text = " | ".join(str(cell) for cell in row)
        ctk.CTkLabel(log_frame, text=text, font=("Arial", 14)).pack(anchor="w", padx=10)

def view_students():
    win = ctk.CTkToplevel(app)
    win.title("Student List")
    win.geometry("500x450")
    ctk.CTkLabel(win, text="📜 Registered Students", font=("Arial", 22, "bold")).pack(pady=10)
    for s in database.get_students():
        ctk.CTkLabel(win, text=f"{s['name']} | {s['contact']} | {s['roll_number']}", font=("Arial", 16)).pack(pady=3)

def update_student():
    win = ctk.CTkToplevel(app)
    win.title("Update Student")
    win.geometry("450x600")

    ctk.CTkLabel(win, text="Enter Roll Number:", font=("Arial", 18)).pack(pady=5)
    roll_entry = ctk.CTkEntry(win, width=350, height=50, font=("Arial", 16))
    roll_entry.pack()

    name_entry = ctk.CTkEntry(win, placeholder_text="New Name", width=350, height=50, font=("Arial", 16))
    contact_entry = ctk.CTkEntry(win, placeholder_text="New Contact", width=350, height=50, font=("Arial", 16))

    def load():
        s = database.get_student_by_roll(roll_entry.get())
        if s:
            name_entry.delete(0, 'end')
            contact_entry.delete(0, 'end')
            name_entry.insert(0, s['name'])
            contact_entry.insert(0, s['contact'])
            name_entry.pack(pady=5)
            contact_entry.pack(pady=5)
        else:
            messagebox.showerror("Not Found", "Student not found!")

    def save():
        ok = database.update_student_details(roll_entry.get(), name_entry.get(), contact_entry.get())
        if ok:
            messagebox.showinfo("✅ Updated", "Student details updated!")
            win.destroy()
        else:
            messagebox.showerror("❌ Error", "Update failed!")

    ctk.CTkButton(win, text="Load Student", command=load, fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=15)
    ctk.CTkButton(win, text="Save Changes", command=save, fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=20)

def delete_student():
    win = ctk.CTkToplevel(app)
    win.title("Delete Student")
    win.geometry("450x300")

    ctk.CTkLabel(win, text="Enter Roll Number:", font=("Arial", 18)).pack(pady=10)
    roll_entry = ctk.CTkEntry(win, width=350, height=50, font=("Arial", 16))
    roll_entry.pack()

    def confirm():
        r = roll_entry.get()
        if messagebox.askyesno("Confirm", f"Delete Roll No: {r}?"):
            if database.delete_student(r):
                messagebox.showinfo("✅ Deleted", "Student deleted.")
                win.destroy()
            else:
                messagebox.showerror("❌ Not Found", "Student not found.")

    ctk.CTkButton(win, text="🗑️ Delete", command=confirm,
                  fg_color="red", hover_color="#D70000", **button_style).pack(pady=25)

def export_data():
    students = database.get_students()
    if not students:
        messagebox.showinfo("Empty", "No data to export.")
        return

    path = asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    if path:
        try:
            with open(path, mode="w", newline='', encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=["name", "contact", "roll_number"])
                writer.writeheader()
                for s in students:
                    s.pop("face_image", None)
                    writer.writerow(s)
            messagebox.showinfo("✅ Exported", f"Exported to {path}")
        except Exception as e:
            messagebox.showerror("❌ Error", str(e))

def register_admin():
    win = ctk.CTkToplevel(app)
    win.title("Register Admin")
    win.geometry("400x400")

    ctk.CTkLabel(win, text="Register New Admin", font=("Arial", 22, "bold")).pack(pady=20)

    uname = ctk.CTkEntry(win, placeholder_text="Username", width=300, height=50, font=("Arial", 16))
    uname.pack(pady=10)
    pwd = ctk.CTkEntry(win, placeholder_text="Password", show="*", width=300, height=50, font=("Arial", 16))
    pwd.pack(pady=10)

    def register():
        if auth.register_admin(uname.get(), pwd.get()):
            messagebox.showinfo("✅ Success", "New admin registered.")
            win.destroy()
        else:
            messagebox.showerror("❌ Failed", "Username already exists or error.")

    ctk.CTkButton(win, text="Register", command=register, fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=30)

# Dashboard
def open_dashboard():
    login_frame.destroy()
    dash = ctk.CTkScrollableFrame(app, corner_radius=20, fg_color="#2E2E2E", width=650, height=700)
    dash.pack(pady=30)
    ctk.CTkLabel(dash, text="📌 Smart Campus Dashboard", font=("Arial", 28, "bold")).pack(pady=25)

    # Section: Student Management
    ctk.CTkLabel(dash, text="👨‍🎓 Student Management", font=("Arial", 22, "bold")).pack(pady=(10, 5))
    for txt, cmd in [
        ("📷 Register Student", register_student),
        ("✏️ Update Student", update_student),
        ("🗑️ Delete Student", delete_student),
        ("📜 View Students", view_students),
        ("🖥️ Train Faces", train_faces)
    ]:
        ctk.CTkButton(dash, text=txt, command=cmd,
                      fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=10)

    # Section: Recognition & Behavior
    ctk.CTkLabel(dash, text="🔍 Recognition & Behavior", font=("Arial", 22, "bold")).pack(pady=(20, 5))
    ctk.CTkButton(dash, text="🔍 Recognize Student", command=recognize_student,
                  fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=10)
    ctk.CTkButton(dash, text="📁 View Misbehavior Logs", command=view_misbehavior_logs,
                  fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=10)

    # Section: Admin Controls
    ctk.CTkLabel(dash, text="🔐 Admin Panel", font=("Arial", 22, "bold")).pack(pady=(20, 5))
    ctk.CTkButton(dash, text="➕ Register New Admin", command=register_admin,
                  fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=10)

    # Section: Reports & Exit
    ctk.CTkLabel(dash, text="📁 Reports & Exit", font=("Arial", 22, "bold")).pack(pady=(20, 5))
    ctk.CTkButton(dash, text="📤 Export to CSV", command=export_data,
                  fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=10)
    ctk.CTkButton(dash, text="❌ Exit", command=app.quit,
                  fg_color="red", hover_color="#D70000", **button_style).pack(pady=15)


# Login Screen
login_frame = ctk.CTkFrame(app, corner_radius=20, fg_color="#2E2E2E", width=500, height=450)
login_frame.pack(pady=170)

ctk.CTkLabel(login_frame, text="🔑 Professor Login", font=("Arial", 26, "bold")).pack(pady=30)
username_entry = ctk.CTkEntry(login_frame, placeholder_text="Username", width=350, height=50, font=("Arial", 18))
username_entry.pack(pady=10)
password_entry = ctk.CTkEntry(login_frame, placeholder_text="Password", show="*", width=350, height=50, font=("Arial", 18))
password_entry.pack(pady=10)

def login():
    if auth.authenticate(username_entry.get(), password_entry.get()):
        messagebox.showinfo("✅ Login Successful", f"Welcome {username_entry.get()}!")
        open_dashboard()
    else:
        messagebox.showerror("❌ Failed", "Invalid credentials.")

ctk.CTkButton(login_frame, text="🔓 Login", command=login,
              fg_color="#3A3A3A", hover_color=hover_color, **button_style).pack(pady=25)

app.mainloop()
