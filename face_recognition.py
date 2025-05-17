import cv2
import numpy as np
import os
import database
import cv2.face  
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Paths
IMAGES_FOLDER = 'images/'
DATASET_FOLDER = 'dataset/'
CASCADE_PATH = os.path.join(os.path.dirname(__file__), "haarcascade", "haarcascade_frontalface_default.xml")

# Load Haar cascade
face_cascade = cv2.CascadeClassifier(CASCADE_PATH)
if face_cascade.empty():
    raise IOError("❌ Error: Haar cascade not loaded. Check the file path.")

# Ensure folders exist
os.makedirs(IMAGES_FOLDER, exist_ok=True)
os.makedirs(DATASET_FOLDER, exist_ok=True)

def capture_face(name, contact, roll_number):
    if not roll_number.isdigit():
        print("❌ Error: Roll number must be numeric!")
        return

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("❌ Error: Cannot access webcam.")
        return

    print("📸 Capturing face... Press 'q' to capture and save face.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("❌ Failed to read from camera.")
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        cv2.imshow("Capture Face - Press Q to Save", frame)

        key = cv2.waitKey(1)
        if key & 0xFF == ord('q'):
            if len(faces) > 0:
                (x, y, w, h) = faces[0]
                filename = f"{roll_number}.jpg"
                filepath = os.path.join(IMAGES_FOLDER, filename)
                face_img = gray[y:y + h, x:x + w]
                cv2.imwrite(filepath, face_img)
                database.add_student(name, contact, roll_number, filename)
                print("✅ Face captured and student added.")
            else:
                print("⚠️ No face detected. Try again.")
            break

    cap.release()
    cv2.destroyAllWindows()

def train_faces():
    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except AttributeError:
        print("❌ Error: OpenCV face module not found. Install opencv-contrib-python.")
        return

    faces, ids = [], []

    students = database.get_students()
    for student in students:
        roll_number = student["roll_number"]
        try:
            img_path = os.path.join(IMAGES_FOLDER, student["face_image"])
            if os.path.exists(img_path):
                image = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                faces.append(image)
                ids.append(int(roll_number))
        except Exception as e:
            print(f"⚠️ Skipping {roll_number}: {e}")

    if faces:
        recognizer.train(faces, np.array(ids))
        recognizer.save(os.path.join(DATASET_FOLDER, 'trainer.yml'))
        print("✅ Training completed.")
    else:
        print("⚠️ No face data found for training.")

def recognize_face():
    model_path = os.path.join(DATASET_FOLDER, 'trainer.yml')
    if not os.path.exists(model_path):
        print("❌ No trained model found.")
        return None

    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read(model_path)
    except AttributeError:
        print("❌ Error: OpenCV face module not found. Install opencv-contrib-python.")
        return None

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("❌ Cannot access webcam.")
        return None

    print("🔍 Looking for face...")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("❌ Failed to read from camera.")
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

        for (x, y, w, h) in faces:
            try:
                face_id, confidence = recognizer.predict(gray[y:y+h, x:x+w])
                student = database.get_student_by_roll(str(face_id))
                if student:
                    print(f"✅ Recognized: {student['name']}")
                    cap.release()
                    cv2.destroyAllWindows()
                    return student
            except:
                continue

        cv2.imshow("Recognizing Face - Press Q to Quit", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    print("⚠️ No face recognized.")
    return None

# -------------------- Misbehavior Logging --------------------

def log_misbehavior(student, reason="Misbehavior"):
    file_path = "misbehavior_log.xlsx"

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Roll Number", "Contact", "Date", "Time", "Reason"])

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    ws.append([student["name"], student["roll_number"], student["contact"], date, time, reason])
    wb.save(file_path)
